"""
crear_planificacion.py
Genera el archivo Excel de planificación StomaSense con todas las tareas
de las Prioridades [3 Meses Mar-May 2026] y [May-Oct | 26].
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, datetime
import os

OUTPUT_FILE = "planificacion_stomasense.xlsx"

# ─────────────────────────────────────────────
# PALETA DE COLORES
# ─────────────────────────────────────────────
C = {
    "header_bg":   "1F3864",   # Azul oscuro encabezado
    "header_font": "FFFFFF",
    "bloque1_bg":  "FFF2CC",   # Amarillo suave  → 3 Meses
    "bloque1_acc": "FFD966",   # Amarillo acento
    "bloque2_bg":  "DDEEFF",   # Azul suave → May-Oct
    "bloque2_acc": "5B9BD5",   # Azul acento
    "proy_bg":     "D9E1F2",   # Fondo fila proyecto
    "subtask_bg":  "FFFFFF",
    "alt_row":     "F2F2F2",
    "green":       "70AD47",
    "yellow":      "FFD966",
    "red":         "FF0000",
    "orange":      "ED7D31",
    "grey":        "A6A6A6",
    "border":      "BFBFBF",
}

STATUS_COLORS = {
    "Pendiente":    C["grey"],
    "En Progreso":  C["bloque2_acc"],
    "Completado":   C["green"],
    "Bloqueado":    "C00000",
    "En Revisión":  C["yellow"],
}

BLOQUE_COLORS = {
    "3 Meses (Mar-May 2026)": C["bloque1_acc"],
    "May-Oct 2026":           C["bloque2_acc"],
}

# ─────────────────────────────────────────────
# DATOS DE TAREAS
# ─────────────────────────────────────────────
TAREAS = [
    # ── BLOQUE 1: 3 Meses ────────────────────────────────────────────
    # Proyecto: Teledetección - Feedback PSH
    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Teledetección – Feedback PSH",
         categoria="Teledetección", responsable="Equipo IA",
         tarea="Issue Review & Diagnóstico",
         desc="Revisión de issues reportados en módulo de teledetección (NDVI+, Hydro Plus, Smart Growth, Weed Detection)",
         inicio=date(2026,3,17), fin=date(2026,3,27), pct=0, estado="En Progreso", notas="Prioridad ALTA"),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Teledetección – Feedback PSH",
         categoria="Teledetección", responsable="Equipo IA",
         tarea="Análisis de resultados por índice",
         desc="Análisis detallado de NDVI+, Hydro Plus, Smart Growth Index y Weed Detection",
         inicio=date(2026,3,28), fin=date(2026,4,7), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Teledetección – Feedback PSH",
         categoria="Teledetección", responsable="Equipo IA + Producto",
         tarea="Preparación reporte Feedback PSH",
         desc="Elaborar reporte ejecutivo con hallazgos, métricas y propuestas de mejora",
         inicio=date(2026,4,8), fin=date(2026,4,18), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Teledetección – Feedback PSH",
         categoria="Teledetección", responsable="Producto",
         tarea="Presentación y validación PSH",
         desc="Sesión de feedback con stakeholders PSH, ajuste de requerimientos",
         inicio=date(2026,4,19), fin=date(2026,4,30), pct=0, estado="Pendiente", notas="Requiere agenda con PSH"),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Teledetección – Feedback PSH",
         categoria="Teledetección", responsable="Equipo IA",
         tarea="Ajustes e implementación post-feedback",
         desc="Correcciones y mejoras identificadas en la sesión PSH",
         inicio=date(2026,5,1), fin=date(2026,5,20), pct=0, estado="Pendiente", notas=""),

    # Proyecto: App Móvil Teledetección E2E
    dict(bloque="3 Meses (Mar-May 2026)", proyecto="App Móvil Teledetección [E2E]",
         categoria="Móvil", responsable="Producto",
         tarea="Definición de requerimientos",
         desc="Levantamiento funcional y técnico de la app móvil end-to-end",
         inicio=date(2026,3,17), fin=date(2026,3,24), pct=0, estado="En Progreso", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="App Móvil Teledetección [E2E]",
         categoria="Móvil", responsable="UX/UI",
         tarea="Diseño UX/UI",
         desc="Wireframes, prototipo y sistema de diseño para la aplicación móvil",
         inicio=date(2026,3,25), fin=date(2026,4,5), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="App Móvil Teledetección [E2E]",
         categoria="Móvil", responsable="Backend",
         tarea="Desarrollo backend – APIs e integraciones",
         desc="Construcción de endpoints, integración con modelos IA de teledetección",
         inicio=date(2026,4,6), fin=date(2026,4,25), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="App Móvil Teledetección [E2E]",
         categoria="Móvil", responsable="Frontend",
         tarea="Desarrollo frontend móvil",
         desc="Implementación de pantallas, consumo de APIs, visualización de índices",
         inicio=date(2026,4,6), fin=date(2026,5,5), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="App Móvil Teledetección [E2E]",
         categoria="Móvil", responsable="QA",
         tarea="Testing & QA",
         desc="Pruebas funcionales, de integración y de campo",
         inicio=date(2026,5,6), fin=date(2026,5,20), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="App Móvil Teledetección [E2E]",
         categoria="Móvil", responsable="DevOps",
         tarea="Despliegue & Go-Live",
         desc="Publicación en stores, despliegue de infraestructura productiva",
         inicio=date(2026,5,21), fin=date(2026,5,31), pct=0, estado="Pendiente", notas=""),

    # Proyecto: Movilidad Empresarial [Modelos | Lab | Operación]
    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Movilidad Empresarial [Modelos|Lab|Op]",
         categoria="Movilidad", responsable="Equipo IA",
         tarea="Modelos – Integración de modelos IA",
         desc="Integración de modelos existentes al flujo de movilidad empresarial",
         inicio=date(2026,3,17), fin=date(2026,4,10), pct=0, estado="En Progreso", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Movilidad Empresarial [Modelos|Lab|Op]",
         categoria="Movilidad", responsable="Laboratorio",
         tarea="Laboratorio – Configuración de ambiente",
         desc="Setup del entorno de laboratorio para pruebas de movilidad",
         inicio=date(2026,4,1), fin=date(2026,4,20), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Movilidad Empresarial [Modelos|Lab|Op]",
         categoria="Movilidad", responsable="Operaciones",
         tarea="Operación – Flujo operativo",
         desc="Configuración y documentación del flujo operativo",
         inicio=date(2026,4,15), fin=date(2026,5,10), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Movilidad Empresarial [Modelos|Lab|Op]",
         categoria="Movilidad", responsable="Tech Lead",
         tarea="Integración Modelos-Lab-Operación",
         desc="Articulación de los tres componentes en pipeline unificado",
         inicio=date(2026,5,11), fin=date(2026,5,25), pct=0, estado="Pendiente", notas=""),

    dict(bloque="3 Meses (Mar-May 2026)", proyecto="Movilidad Empresarial [Modelos|Lab|Op]",
         categoria="Movilidad", responsable="QA",
         tarea="Pruebas end-to-end",
         desc="Validación integral del flujo completo",
         inicio=date(2026,5,26), fin=date(2026,5,31), pct=0, estado="Pendiente", notas=""),

    # ── BLOQUE 2: May-Oct 2026 ────────────────────────────────────────
    # Proyecto: Modelos IA Agronomía
    dict(bloque="May-Oct 2026", proyecto="Modelos IA – Pol|Maduración|TCH|Caña",
         categoria="Modelos IA", responsable="Equipo IA",
         tarea="Recopilación y limpieza de datos",
         desc="Consolidación de datasets históricos para entrenamiento de modelos",
         inicio=date(2026,5,1), fin=date(2026,5,20), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Modelos IA – Pol|Maduración|TCH|Caña",
         categoria="Modelos IA", responsable="Data Science",
         tarea="Entrenamiento modelo Pol (polarimetría)",
         desc="Fejuste y entrenamiento del modelo de Pol en caña",
         inicio=date(2026,5,21), fin=date(2026,6,15), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Modelos IA – Pol|Maduración|TCH|Caña",
         categoria="Modelos IA", responsable="Data Science",
         tarea="Entrenamiento modelo Maduración",
         desc="Modelo predictivo de maduración de caña",
         inicio=date(2026,6,1), fin=date(2026,6,30), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Modelos IA – Pol|Maduración|TCH|Caña",
         categoria="Modelos IA", responsable="Data Science",
         tarea="Entrenamiento modelo TCH (Toneladas Caña/Ha)",
         desc="Predicción de rendimiento TCH",
         inicio=date(2026,6,15), fin=date(2026,7,15), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Modelos IA – Pol|Maduración|TCH|Caña",
         categoria="Modelos IA", responsable="Data Science",
         tarea="Modelo Caña / No Caña (clasificación)",
         desc="Clasificador binario de presencia de caña en imágenes satelitales",
         inicio=date(2026,6,1), fin=date(2026,7,15), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Modelos IA – Pol|Maduración|TCH|Caña",
         categoria="Modelos IA", responsable="MLOps",
         tarea="Validación y ajuste de modelos",
         desc="Validación cruzada, métricas en campo y ajuste de hiperparámetros",
         inicio=date(2026,7,16), fin=date(2026,8,15), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Modelos IA – Pol|Maduración|TCH|Caña",
         categoria="Modelos IA", responsable="Backend",
         tarea="Integración a plataforma",
         desc="Despliegue de modelos como microservicios en la plataforma",
         inicio=date(2026,8,16), fin=date(2026,9,15), pct=0, estado="Pendiente", notas=""),

    # Proyecto: Frescura de Caña
    dict(bloque="May-Oct 2026", proyecto="Frescura de Caña",
         categoria="Modelos IA", responsable="Data Science",
         tarea="Diseño del modelo de frescura",
         desc="Definición de features, metodología y métricas del modelo",
         inicio=date(2026,5,1), fin=date(2026,5,31), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Frescura de Caña",
         categoria="Modelos IA", responsable="Campo",
         tarea="Recolección de datos sensoriales",
         desc="Captura de datos en campo (imágenes, sensores, muestras)",
         inicio=date(2026,6,1), fin=date(2026,6,30), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Frescura de Caña",
         categoria="Modelos IA", responsable="Data Science",
         tarea="Entrenamiento modelo frescura",
         desc="Entrenamiento y evaluación del modelo",
         inicio=date(2026,7,1), fin=date(2026,7,31), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Frescura de Caña",
         categoria="Modelos IA", responsable="QA + Campo",
         tarea="Validación en campo",
         desc="Pruebas reales y comparación con método convencional",
         inicio=date(2026,8,1), fin=date(2026,8,31), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Frescura de Caña",
         categoria="Modelos IA", responsable="DevOps",
         tarea="Despliegue productivo",
         desc="Publicación del modelo en producción",
         inicio=date(2026,9,1), fin=date(2026,9,30), pct=0, estado="Pendiente", notas=""),

    # Proyecto: Salud Financiera | Rentabilidad Fincas
    dict(bloque="May-Oct 2026", proyecto="Salud Financiera | Rentabilidad Fincas",
         categoria="Modelos IA", responsable="Negocio + IA",
         tarea="Levantamiento de indicadores financieros",
         desc="Definición de KPIs de rentabilidad y salud financiera por finca",
         inicio=date(2026,5,15), fin=date(2026,6,15), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Salud Financiera | Rentabilidad Fincas",
         categoria="Modelos IA", responsable="Backend",
         tarea="Integración de datos financieros",
         desc="Conexión con fuentes de datos contables/ERP",
         inicio=date(2026,6,16), fin=date(2026,7,15), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Salud Financiera | Rentabilidad Fincas",
         categoria="Modelos IA", responsable="Data Science",
         tarea="Desarrollo modelos predictivos financieros",
         desc="Modelos de predicción de rentabilidad y alertas de salud financiera",
         inicio=date(2026,7,16), fin=date(2026,8,31), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Salud Financiera | Rentabilidad Fincas",
         categoria="Modelos IA", responsable="Frontend",
         tarea="Dashboard de rentabilidad y salud financiera",
         desc="Visualización interactiva de indicadores financieros por finca",
         inicio=date(2026,9,1), fin=date(2026,10,15), pct=0, estado="Pendiente", notas=""),

    # Proyecto: IoT - Telemetría [Riesgos]
    dict(bloque="May-Oct 2026", proyecto="IoT Analytics – Telemetría [Riesgos]",
         categoria="IoT Analytics", responsable="IoT Team",
         tarea="Definición de alertas y umbrales de riesgo",
         desc="Catálogo de riesgos, umbrales y lógica de alertas IoT",
         inicio=date(2026,5,15), fin=date(2026,6,1), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="IoT Analytics – Telemetría [Riesgos]",
         categoria="IoT Analytics", responsable="IoT Team",
         tarea="Integración de sensores IoT",
         desc="Conexión y configuración de sensores de riego, clima, suelo",
         inicio=date(2026,6,1), fin=date(2026,6,30), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="IoT Analytics – Telemetría [Riesgos]",
         categoria="IoT Analytics", responsable="Frontend",
         tarea="Desarrollo dashboard de telemetría",
         desc="Panel en tiempo real de métricas y alertas de riesgo",
         inicio=date(2026,7,1), fin=date(2026,7,31), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="IoT Analytics – Telemetría [Riesgos]",
         categoria="IoT Analytics", responsable="QA + Campo",
         tarea="Testing y calibración de sensores",
         desc="Validación de precisión y calibración en campo",
         inicio=date(2026,8,1), fin=date(2026,8,31), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="IoT Analytics – Telemetría [Riesgos]",
         categoria="IoT Analytics", responsable="DevOps",
         tarea="Despliegue productivo telemetría",
         desc="Go-live del sistema de telemetría con alertas en producción",
         inicio=date(2026,9,1), fin=date(2026,9,30), pct=0, estado="Pendiente", notas=""),

    # Proyecto: IoT - Movilidad Empresarial [Avance Labores]
    dict(bloque="May-Oct 2026", proyecto="IoT – Movilidad Empresarial [Avance Labores]",
         categoria="IoT Analytics", responsable="Producto",
         tarea="Diseño del flujo de avance de labores",
         desc="Mapeo del proceso de reporte de avance de labores en campo",
         inicio=date(2026,6,1), fin=date(2026,6,20), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="IoT – Movilidad Empresarial [Avance Labores]",
         categoria="IoT Analytics", responsable="Backend",
         tarea="Integración con módulos existentes",
         desc="Conexión con módulos operacionales y de movilidad ya desplegados",
         inicio=date(2026,6,21), fin=date(2026,7,20), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="IoT – Movilidad Empresarial [Avance Labores]",
         categoria="IoT Analytics", responsable="QA + Usuarios",
         tarea="Pruebas con usuarios (UAT)",
         desc="Sesiones de prueba con usuarios de campo, iteraciones",
         inicio=date(2026,7,21), fin=date(2026,8,20), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="IoT – Movilidad Empresarial [Avance Labores]",
         categoria="IoT Analytics", responsable="DevOps",
         tarea="Ajustes finales y despliegue",
         desc="Correcciones post-UAT y go-live",
         inicio=date(2026,8,21), fin=date(2026,9,30), pct=0, estado="Pendiente", notas=""),

    # Proyecto: Módulo de Riego
    dict(bloque="May-Oct 2026", proyecto="Módulo Operacional – Módulo de Riego",
         categoria="Módulo Operacional", responsable="Producto + IoT",
         tarea="Requerimientos y diseño funcional",
         desc="Especificación funcional y técnica del módulo de riego",
         inicio=date(2026,5,1), fin=date(2026,5,25), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Módulo Operacional – Módulo de Riego",
         categoria="Módulo Operacional", responsable="Backend + IoT",
         tarea="Desarrollo core del módulo",
         desc="Implementación de lógica de negocio, algoritmos de riego",
         inicio=date(2026,5,26), fin=date(2026,7,15), pct=0, estado="Pendiente", notas="Dependencia: telemetría"),

    dict(bloque="May-Oct 2026", proyecto="Módulo Operacional – Módulo de Riego",
         categoria="Módulo Operacional", responsable="IoT Team",
         tarea="Integración con telemetría de sensores",
         desc="Conexión con sensores de humedad, clima y actuadores de riego",
         inicio=date(2026,7,16), fin=date(2026,8,15), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Módulo Operacional – Módulo de Riego",
         categoria="Módulo Operacional", responsable="QA + Campo",
         tarea="Testing en campo",
         desc="Pruebas reales en fincas piloto",
         inicio=date(2026,8,16), fin=date(2026,9,15), pct=0, estado="Pendiente", notas=""),

    dict(bloque="May-Oct 2026", proyecto="Módulo Operacional – Módulo de Riego",
         categoria="Módulo Operacional", responsable="Operaciones",
         tarea="Despliegue y capacitación",
         desc="Go-live, documentación y capacitación a usuarios finales",
         inicio=date(2026,9,16), fin=date(2026,10,15), pct=0, estado="Pendiente", notas=""),
]

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(color=C["border"]):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font():
    return Font(name="Calibri", bold=True, color=C["header_font"], size=10)

def normal_font(bold=False, color="000000", size=9):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def add_border_to_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border()

# ─────────────────────────────────────────────
# CREAR EXCEL
# ─────────────────────────────────────────────
def crear_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planificación"

    # ── Fila de título principal ──
    ws.merge_cells("A1:N1")
    tc = ws["A1"]
    tc.value = "STOMASENSE – ROADMAP DE PRIORIDADES 2026"
    tc.fill = fill(C["header_bg"])
    tc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    tc.alignment = center()
    ws.row_dimensions[1].height = 32

    # ── Subtítulo ──
    ws.merge_cells("A2:N2")
    sc = ws["A2"]
    sc.value = f"Generado: {datetime.today().strftime('%d/%m/%Y')}  |  Bloques: [3 Meses Mar-May 2026] y [May-Oct 2026]"
    sc.fill = fill("2F5496")
    sc.font = Font(name="Calibri", italic=True, color="FFFFFF", size=10)
    sc.alignment = center()
    ws.row_dimensions[2].height = 18

    # ── Encabezados de columnas ──
    COLS = [
        ("ID",           6),
        ("Bloque",       20),
        ("Proyecto",     32),
        ("Categoría",    20),
        ("Tarea",        38),
        ("Descripción",  50),
        ("Responsable",  20),
        ("Inicio",       13),
        ("Fin",          13),
        ("Días",          7),
        ("% Avance",     10),
        ("Estado",       14),
        ("Notas",        30),
        ("Actualizado",  14),
    ]

    for col_idx, (col_name, col_w) in enumerate(COLS, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = col_name
        cell.fill = fill(C["header_bg"])
        cell.font = header_font()
        cell.alignment = center()
        cell.border = border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w

    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    # ── Data Validation – Estado ──
    dv_status = DataValidation(
        type="list",
        formula1='"Pendiente,En Progreso,Completado,Bloqueado,En Revisión"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_status.sqref = f"L4:L{3 + len(TAREAS)}"
    ws.add_data_validation(dv_status)

    # ── Data Validation – % Avance (0-100) ──
    dv_pct = DataValidation(
        type="whole", operator="between", formula1="0", formula2="100",
        allow_blank=False,
    )
    dv_pct.sqref = f"K4:K{3 + len(TAREAS)}"
    ws.add_data_validation(dv_pct)

    # ── Escribir filas ──
    prev_bloque = None
    prev_proyecto = None

    for idx, t in enumerate(TAREAS, start=1):
        row = idx + 3
        duracion = (t["fin"] - t["inicio"]).days + 1

        # Color de fondo
        if t["bloque"] == "3 Meses (Mar-May 2026)":
            row_bg = C["bloque1_bg"] if idx % 2 == 0 else "FFFDE7"
        else:
            row_bg = C["bloque2_bg"] if idx % 2 == 0 else "EEF4FB"

        values = [
            idx,
            t["bloque"],
            t["proyecto"],
            t["categoria"],
            t["tarea"],
            t["desc"],
            t["responsable"],
            t["inicio"],
            t["fin"],
            duracion,
            t["pct"],
            t["estado"],
            t["notas"],
            datetime.today().strftime("%d/%m/%Y"),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = fill(row_bg)
            cell.border = border()
            cell.font = normal_font()

            if col_idx in (1, 10, 11):  # ID, Días, %
                cell.alignment = center()
            elif col_idx in (8, 9, 14):  # fechas
                cell.alignment = center()
                if isinstance(val, date):
                    cell.number_format = "DD/MM/YYYY"
            else:
                cell.alignment = left()

        # Colorear Estado
        estado_cell = ws.cell(row=row, column=12)
        st_color = STATUS_COLORS.get(t["estado"], C["grey"])
        estado_cell.fill = fill(st_color)
        if t["estado"] == "Completado":
            estado_cell.font = normal_font(bold=True, color="FFFFFF")
        elif t["estado"] in ("En Progreso",):
            estado_cell.font = normal_font(bold=True, color="FFFFFF")
        else:
            estado_cell.font = normal_font(bold=False, color="000000")

        # Separador visual entre bloques
        if prev_bloque and prev_bloque != t["bloque"]:
            for c in range(1, 15):
                ws.cell(row=row, column=c).border = Border(
                    top=Side(style="medium", color="1F3864"),
                    left=Side(style="thin", color=C["border"]),
                    right=Side(style="thin", color=C["border"]),
                    bottom=Side(style="thin", color=C["border"]),
                )

        # Separador visual entre proyectos
        elif prev_proyecto and prev_proyecto != t["proyecto"]:
            for c in range(1, 15):
                existing = ws.cell(row=row, column=c).border
                ws.cell(row=row, column=c).border = Border(
                    top=Side(style="medium", color=C["bloque1_acc"] if t["bloque"] == "3 Meses (Mar-May 2026)" else C["bloque2_acc"]),
                    left=Side(style="thin", color=C["border"]),
                    right=Side(style="thin", color=C["border"]),
                    bottom=Side(style="thin", color=C["border"]),
                )

        prev_bloque = t["bloque"]
        prev_proyecto = t["proyecto"]

        ws.row_dimensions[row].height = 28

    # ── Hoja de leyenda ──
    ws_leyenda = wb.create_sheet("Leyenda")
    leyenda_data = [
        ("COLOR ESTADO", "SIGNIFICADO"),
        ("Pendiente", "Tarea no iniciada"),
        ("En Progreso", "Tarea en ejecución activa"),
        ("Completado", "Tarea finalizada"),
        ("Bloqueado", "Tarea detenida por impedimento"),
        ("En Revisión", "Tarea en fase de revisión/aprobación"),
    ]
    ws_leyenda.column_dimensions["A"].width = 20
    ws_leyenda.column_dimensions["B"].width = 35
    for r, (k, v) in enumerate(leyenda_data, start=1):
        ca = ws_leyenda.cell(row=r, column=1, value=k)
        cb = ws_leyenda.cell(row=r, column=2, value=v)
        if r == 1:
            ca.fill = fill(C["header_bg"]); ca.font = header_font()
            cb.fill = fill(C["header_bg"]); cb.font = header_font()
        else:
            color = list(STATUS_COLORS.values())[r - 2]
            ca.fill = fill(color)
            ca.font = normal_font(bold=True, color="FFFFFF" if r > 2 else "000000")
            cb.font = normal_font()
        ca.border = border(); cb.border = border()
        ca.alignment = center(); cb.alignment = left()

    # ── Hoja Instrucciones ──
    ws_inst = wb.create_sheet("Instrucciones")
    instrucciones = [
        "=== INSTRUCCIONES DE USO ===",
        "",
        "1. Actualiza el % Avance (columna K) con un valor entre 0 y 100.",
        "2. Cambia el Estado (columna L) usando el menú desplegable.",
        "3. Puedes modificar las fechas de Inicio y Fin (columnas H e I).",
        "4. Agrega notas en la columna M.",
        "5. Para regenerar el Gantt interactivo, ejecuta:",
        "   > python generar_gantt.py",
        "   Esto creará/actualizará el archivo 'gantt_stomasense.html'.",
        "",
        "=== PROYECTOS ===",
        "Bloque 3 Meses (Mar-May 2026):",
        "  - Teledetección – Feedback PSH",
        "  - App Móvil Teledetección [E2E]",
        "  - Movilidad Empresarial [Modelos|Lab|Op]",
        "",
        "Bloque May-Oct 2026:",
        "  - Modelos IA – Pol|Maduración|TCH|Caña",
        "  - Frescura de Caña",
        "  - Salud Financiera | Rentabilidad Fincas",
        "  - IoT Analytics – Telemetría [Riesgos]",
        "  - IoT – Movilidad Empresarial [Avance Labores]",
        "  - Módulo Operacional – Módulo de Riego",
    ]
    ws_inst.column_dimensions["A"].width = 65
    for r, txt in enumerate(instrucciones, start=1):
        cell = ws_inst.cell(row=r, column=1, value=txt)
        cell.font = Font(name="Calibri", size=10,
                         bold=txt.startswith("==="),
                         color="1F3864" if txt.startswith("===") else "000000")
        cell.alignment = left()

    wb.save(OUTPUT_FILE)
    print(f"✅  Excel guardado: {OUTPUT_FILE}")
    print(f"   Total tareas: {len(TAREAS)}")
    proyectos = sorted(set(t["proyecto"] for t in TAREAS))
    for p in proyectos:
        n = sum(1 for t in TAREAS if t["proyecto"] == p)
        print(f"   · {p}: {n} tareas")


if __name__ == "__main__":
    crear_excel()
